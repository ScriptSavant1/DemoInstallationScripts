"""
generate_soe.py
Generates LRE 25.1 -> 26.1 Upgrade  Sequence of Events (SOE) workbook.
Run:  python generate_soe.py
Output: LRE_Upgrade_SOE.xlsx  (in the same directory)
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date

# ── Colour palette ────────────────────────────────────────────────────────────
CLR = {
    "navy":        "1F3864",   # dark header background
    "blue":        "2E75B6",   # section header
    "light_blue":  "BDD7EE",   # alternate row
    "green":       "375623",   # success / done
    "green_bg":    "E2EFDA",   # light green fill
    "amber":       "9C5700",   # warning text
    "amber_bg":    "FFEB9C",   # warning fill
    "red":         "9C0006",   # critical / backout
    "red_bg":      "FFC7CE",   # critical fill
    "gray":        "808080",
    "light_gray":  "F2F2F2",
    "white":       "FFFFFF",
    "yellow_hdr":  "FFD966",   # cover page accent
    "teal":        "1F6B75",   # sheet accent
    "teal_light":  "D9EFF0",
}

# ── Helper – thin border ──────────────────────────────────────────────────────
def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border():
    s = Side(style="medium", color="1F3864")
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ── Helper – write a styled cell ─────────────────────────────────────────────
def wc(ws, row, col, value,
       bold=False, fg="000000", bg=None, sz=10,
       h_align="left", italic=False, wrap=True, border=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = font(bold=bold, color=fg, size=sz, italic=italic)
    cell.alignment = align(h=h_align, wrap=wrap)
    if bg:
        cell.fill = fill(bg)
    if border:
        cell.border = thin_border()
    return cell

# ── Helper – section header spanning columns ──────────────────────────────────
def section_hdr(ws, row, start_col, end_col, text, bg=CLR["blue"]):
    ws.merge_cells(start_row=row, start_column=start_col,
                   end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=start_col, value=text)
    c.fill      = fill(bg)
    c.font      = font(bold=True, color=CLR["white"], size=11)
    c.alignment = align(h="left")
    for col in range(start_col, end_col + 1):
        ws.cell(row=row, column=col).border = thin_border()

# ── Helper – column header row ────────────────────────────────────────────────
def col_headers(ws, row, headers, bg=CLR["navy"]):
    for col, text in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=text)
        c.fill      = fill(bg)
        c.font      = font(bold=True, color=CLR["white"], size=10)
        c.alignment = align(h="center")
        c.border    = thin_border()

# ── Helper – freeze panes after header rows ───────────────────────────────────
def freeze(ws, cell="A3"):
    ws.freeze_panes = cell

# ── Helper – auto-width (approximate) ────────────────────────────────────────
def set_col_widths(ws, widths):
    """widths: list of (col_letter_or_idx, width)"""
    for col, w in widths:
        if isinstance(col, int):
            col = get_column_letter(col)
        ws.column_dimensions[col].width = w

# ── Helper – data row with alternating shading ───────────────────────────────
def data_row(ws, row, values, alt=False, bold_col=None, status_col=None):
    """
    values: list aligned to columns starting at 1.
    alt: True → light shading
    bold_col: 1-based column index to make bold
    status_col: 1-based column index that contains a status keyword
                (auto-colours the cell)
    """
    bg_base = CLR["light_gray"] if alt else CLR["white"]
    for col, val in enumerate(values, 1):
        bg = bg_base
        fg = "000000"
        bd = False
        bld = (col == bold_col)

        if col == status_col and isinstance(val, str):
            v = val.upper()
            if v in ("DONE", "COMPLETE", "PASS"):
                bg, fg = CLR["green_bg"], CLR["green"]
                bld = True
            elif v in ("PENDING", "NOT STARTED"):
                bg, fg = CLR["light_gray"], CLR["gray"]
            elif v in ("IN PROGRESS",):
                bg, fg = CLR["amber_bg"], CLR["amber"]
                bld = True
            elif v in ("FAIL", "FAILED", "ERROR"):
                bg, fg = CLR["red_bg"], CLR["red"]
                bld = True

        c = ws.cell(row=row, column=col, value=val)
        c.font      = font(bold=bld, color=fg, size=10)
        c.alignment = align()
        c.fill      = fill(bg)
        c.border    = thin_border()


# ==============================================================================
# SHEET 1 — COVER
# ==============================================================================
def build_cover(wb):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height  = 6
    ws.row_dimensions[2].height  = 50
    ws.row_dimensions[3].height  = 30
    ws.row_dimensions[4].height  = 22
    ws.row_dimensions[5].height  = 22
    ws.row_dimensions[6].height  = 22
    ws.row_dimensions[7].height  = 22
    ws.row_dimensions[8].height  = 22
    ws.row_dimensions[9].height  = 18
    ws.row_dimensions[10].height = 30

    set_col_widths(ws, [("A", 4), ("B", 28), ("C", 42), ("D", 22),
                        ("E", 22), ("F", 4)])

    # Title banner
    ws.merge_cells("B2:E2")
    c = ws["B2"]
    c.value     = "OpenText LRE Upgrade  |  25.1 → 26.1"
    c.fill      = fill(CLR["navy"])
    c.font      = Font(bold=True, color=CLR["yellow_hdr"], size=22, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B3:E3")
    c = ws["B3"]
    c.value     = "Sequence of Events (SOE)  —  Installation & Upgrade Runbook"
    c.fill      = fill(CLR["blue"])
    c.font      = Font(bold=True, color=CLR["white"], size=14, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Meta table
    meta = [
        ("Document Date",      str(date.today().strftime("%d %B %Y"))),
        ("Upgrade Version",    "25.1  →  26.1"),
        ("Prepared By",        "<Your Name / Team>"),
        ("Approved By",        "<Manager / Change Manager>"),
        ("Change Ref #",       "<ITSM Change # / CR-XXXXX>"),
    ]
    for i, (label, value) in enumerate(meta, 4):
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=2)
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=5)
        wc(ws, i, 2, label, bold=True, fg=CLR["white"], bg=CLR["navy"], sz=10,
           border=True)
        wc(ws, i, 3, value, bg=CLR["light_gray"], sz=10, border=True)

    # Workbook guide
    ws.merge_cells("B10:E10")
    c = ws["B10"]
    c.value     = "Workbook Contents"
    c.fill      = fill(CLR["blue"])
    c.font      = Font(bold=True, color=CLR["white"], size=11, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")

    guide = [
        ("Prerequisites",         "Hardware, software, network, and account requirements before starting."),
        ("Pre-Upgrade Checks",    "Run diagnostic script on every machine. Confirm all checks PASS."),
        ("LRE Server Upgrade",    "Step-by-step guide to upgrade the LRE Server (script 01)."),
        ("Host Upgrades",         "Upgrade Controller, Data Processor, and Load Generator hosts (scripts 02–04)."),
        ("Post-Upgrade Verify",   "Health checks and browser tests to confirm successful upgrade (script 05)."),
        ("Backout Plan",          "Step-by-step rollback procedure if the upgrade must be reversed."),
    ]
    for i, (sheet, desc) in enumerate(guide, 11):
        ws.row_dimensions[i].height = 20
        wc(ws, i, 2, sheet, bold=True, bg=CLR["light_blue"], sz=10, border=True)
        ws.merge_cells(start_row=i, start_column=3,
                       end_row=i,   end_column=5)
        wc(ws, i, 3, desc, bg=CLR["white"], sz=10, border=True)

    # Disclaimer row
    r = 11 + len(guide) + 1
    ws.row_dimensions[r].height = 28
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    wc(ws, r, 2,
       "⚠  Always obtain Change Advisory Board (CAB) approval before executing this runbook in production.",
       bold=True, fg=CLR["amber"], bg=CLR["amber_bg"], sz=10, border=True)


# ==============================================================================
# SHEET 2 — PREREQUISITES
# ==============================================================================
def build_prerequisites(wb):
    ws = wb.create_sheet("Prerequisites")
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, [("A", 5), ("B", 6), ("C", 32), ("D", 52), ("E", 18), ("F", 5)])

    # Title
    ws.merge_cells("B1:E1")
    c = ws["B1"]
    c.value     = "PREREQUISITES  —  LRE 25.1 → 26.1 Upgrade"
    c.fill      = fill(CLR["navy"])
    c.font      = Font(bold=True, color=CLR["yellow_hdr"], size=14, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["#", "Category", "Requirement", "Detail / Acceptance Criteria", "Status"]
    col_headers(ws, 2, [""] + headers, bg=CLR["navy"])
    ws.row_dimensions[2].height = 22
    freeze(ws, "B3")

    rows = [
        # ---- Access & Accounts ----
        ("ACCESS & ACCOUNTS", None, None, None),
        (1,  "Access & Accounts",  "Local Administrator rights",
         "Must run all PowerShell upgrade scripts under a local admin account. Right-click → 'Run as Administrator'.",
         "Pending"),
        (2,  "Access & Accounts",  "Network share access (read)",
         r"Account must have read access to \\fileserver\LRE_26.1 from every upgrade target machine.",
         "Pending"),
        (3,  "Access & Accounts",  "DB Admin credentials",
         "SQL Server 'sa' (or equivalent DBA) credentials required. Kept in upgrade_config.ps1 or entered at prompt.",
         "Pending"),
        (4,  "Access & Accounts",  "DB Application user credentials",
         "lre_user (or equivalent) credentials required for LRE databases.",
         "Pending"),
        (5,  "Access & Accounts",  "LRE System user credentials",
         "IUSR_METRO (or custom domain user) password required. Must match on ALL nodes.",
         "Pending"),

        # ---- Installer Media ----
        ("INSTALLER MEDIA", None, None, None),
        (6,  "Installer Media",    "26.1 installer share populated",
         r"\\fileserver\LRE_26.1\Setup\En\setup_server.exe and setup_host.exe must exist.",
         "Pending"),
        (7,  "Installer Media",    "OneLG installer present",
         r"\\fileserver\LRE_26.1\Standalone Applications\SetupOneLG.exe must exist.",
         "Pending"),
        (8,  "Installer Media",    "Prerequisites in share",
         r"Share must contain: dotnet48\ndp48-*.exe, dotnet_hosting\dotnet-hosting-8.0.*.exe, vc2022_redist_x86\vc_redist.x86.exe, vc2022_redist_x64\vc_redist.x64.exe",
         "Pending"),
        (9,  "Installer Media",    "Vendor UserInput.xml present",
         r"\\fileserver\LRE_26.1\Setup\Install\Server\UserInput.xml  (merge template used by script 01).",
         "Pending"),

        # ---- Software & OS ----
        ("SOFTWARE & OS", None, None, None),
        (10, "Software & OS",      "Current version confirmed = 25.1",
         "Run 00_PreUpgradeChecks.ps1 on every machine. 'LRE Version is 25.1' check must PASS.",
         "Pending"),
        (11, "Software & OS",      "Windows OS (English locale)",
         "All upgrade targets must be Windows Server with en-* locale. Non-English locales may cause silent installer issues.",
         "Pending"),
        (12, "Software & OS",      "IIS installed (LRE Server only)",
         "W3SVC service must be present on the LRE Server machine before running script 01.",
         "Pending"),
        (13, "Software & OS",      "No pending Windows reboots",
         "Run 00_PreUpgradeChecks.ps1 — 'No Pending Reboot' must PASS. Reboot and re-run if needed.",
         "Pending"),

        # ---- Disk & Network ----
        ("DISK & NETWORK", None, None, None),
        (14, "Disk & Network",     "≥ 20 GB free disk space on C:\\",
         "Each target machine (Server, Controller, DP, LG) must have at least 20 GB free on C:\\.",
         "Pending"),
        (15, "Disk & Network",     "C:\\LRE_UpgradeLogs directory writable",
         "Script creates C:\\LRE_UpgradeLogs automatically. Confirm the account has write access to C:\\.",
         "Pending"),
        (16, "Disk & Network",     "All machines can reach installer share",
         r"Test from each machine: Test-Path '\\fileserver\LRE_26.1' should return True.",
         "Pending"),
        (17, "Disk & Network",     "LRE Server FQDN resolves",
         "lre.yourdomain.com (or your IisSecureHostName) must resolve from all host machines.",
         "Pending"),

        # ---- Database ----
        ("DATABASE", None, None, None),
        (18, "Database",           "Database backups taken",
         "Confirm DBA has backed up: LRE_LAB, LRE_ADMIN, LRE_SITE databases BEFORE starting upgrade.",
         "Pending"),
        (19, "Database",           "SQL Server accessible",
         "sql-server-01:1433 (or your DbServerHost:DbServerPort) must be reachable from the LRE Server.",
         "Pending"),

        # ---- Security ----
        ("SECURITY & ENCRYPTION", None, None, None),
        (20, "Security",           "Secure Passphrase agreed (≥ 12 chars)",
         "A single passphrase must be agreed upon BEFORE upgrade. It must be used identically on ALL nodes.",
         "Pending"),
        (21, "Security",           "Encryption key (LW_CRYPTO_INIT_STRING) known",
         "CryptoKey value must be identical across Server, Controller, DP, and LG. Obtain from current pcs.config if needed.",
         "Pending"),
        (22, "Security",           "TLS certificate available",
         "A valid SSL certificate for lre.yourdomain.com must be in the Windows Certificate Store (or .pfx file ready for import).",
         "Pending"),

        # ---- Change Management ----
        ("CHANGE MANAGEMENT", None, None, None),
        (23, "Change Management",  "CAB approval obtained",
         "A Change Request must be approved by the Change Advisory Board before executing in production.",
         "Pending"),
        (24, "Change Management",  "Maintenance window scheduled",
         "Coordinate downtime window with all stakeholders. Estimate: 2–4 hours for a single-server environment.",
         "Pending"),
        (25, "Change Management",  "Rollback decision point agreed",
         "Agree in advance: at what stage will the team decide to roll back? (Recommendation: if server upgrade fails at Phase 7.)",
         "Pending"),
    ]

    r = 3
    alt = False
    for item in rows:
        if item[1] is None:
            # Section header
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            section_hdr(ws, r, 2, 6, f"  {item[0]}")
            ws.row_dimensions[r].height = 20
            r += 1
            alt = False
            continue

        num, cat, req, detail, status = item
        bg = CLR["light_gray"] if alt else CLR["white"]
        wc(ws, r, 2, num, bold=True, bg=bg, h_align="center", border=True)
        wc(ws, r, 3, cat, bg=bg, border=True, italic=True, fg=CLR["gray"])
        wc(ws, r, 4, req, bold=True, bg=bg, border=True)
        wc(ws, r, 5, detail, bg=bg, border=True)

        # Status cell with colour
        s_bg = CLR["amber_bg"]
        s_fg = CLR["amber"]
        wc(ws, r, 6, status, bold=True, fg=s_fg, bg=s_bg,
           h_align="center", border=True)

        ws.row_dimensions[r].height = 32
        r += 1
        alt = not alt


# ==============================================================================
# SHEET 3 — PRE-UPGRADE CHECKS
# ==============================================================================
def build_pre_checks(wb):
    ws = wb.create_sheet("Pre-Upgrade Checks")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 5), ("B", 5), ("C", 30), ("D", 50),
                        ("E", 22), ("F", 18), ("G", 5)])

    ws.merge_cells("B1:F1")
    c = ws["B1"]
    c.value     = "PRE-UPGRADE CHECKS  —  Run on EVERY machine before upgrade"
    c.fill      = fill(CLR["navy"])
    c.font      = Font(bold=True, color=CLR["yellow_hdr"], size=14, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    col_headers(ws, 2, ["", "#", "Check", "How to Verify / Expected Result",
                         "Machine(s)", "Result"], bg=CLR["navy"])
    ws.row_dimensions[2].height = 22
    freeze(ws, "B3")

    note_row = 3
    ws.merge_cells(f"B{note_row}:F{note_row}")
    wc(ws, note_row, 2,
       "▶  Script:  .\\00_PreUpgradeChecks.ps1  (run elevated on each machine — Server, Controller, Data Processor, Load Generator)",
       bold=True, fg=CLR["teal"], bg=CLR["teal_light"], sz=10, border=True)
    ws.row_dimensions[note_row].height = 22

    checks = [
        (1,  "Administrator Privileges",
         "Run script elevated (right-click → Run as Administrator). Script self-checks and reports PASS/FAIL.",
         "All machines", "Pending"),
        (2,  "Operating System",
         "Script reports Windows OS and build number. Must be Windows Server.",
         "All machines", "Pending"),
        (3,  "System Locale = English",
         "Locale must show 'en-*'. Non-English may cause installer issues.",
         "All machines", "Pending"),
        (4,  "No Pending Reboot",
         "Must show 'No pending reboot'. If FAIL: reboot the machine and re-run checks.",
         "All machines", "Pending"),
        (5,  "Installed LRE Version = 25.1",
         "Registry check confirms current installation is 25.1. Script exits with error if not 25.1.",
         "All machines", "Pending"),
        (6,  "Disk Space ≥ 20 GB free on C:\\",
         "Must show PASS. Free up disk space if FAIL before proceeding.",
         "All machines", "Pending"),
        (7,  ".NET Framework 4.8",
         "Registry Release value must be ≥ 528040. Script will install it in Phase 5 if missing.",
         "All machines", "Pending"),
        (8,  ".NET Core Hosting 8.x",
         "dotnet --list-runtimes shows Microsoft.AspNetCore.App 8.x. Script installs if missing.",
         "All machines", "Pending"),
        (9,  "IIS Installed (W3SVC)",
         "W3SVC service must be present. Mandatory for LRE Server machine only.",
         "LRE Server", "Pending"),
        (10, "LRE Services Status",
         "Script reports current state of IIS, Backend, Alerts, Agent, Remote Mgmt services. Note down states.",
         "All machines", "Pending"),
        (11, "Installer Share Accessible",
         r"\\fileserver\LRE_26.1 must return True. If FAIL: check share path and permissions.",
         "All machines", "Pending"),
        (12, "setup_server.exe / setup_host.exe / SetupOneLG.exe found",
         "All three installer executables must be found. If FAIL: verify installer media in the share.",
         "All machines", "Pending"),
        (13, "No Installer Processes Running",
         "No msiexec / setup processes must be running. Kill any stale installer processes if FAIL.",
         "All machines", "Pending"),
        (14, "Database Backups Confirmed",
         "Manual step — contact DBA and get written confirmation that LRE_LAB, LRE_ADMIN, LRE_SITE are backed up.",
         "LRE Server", "Pending"),
        (15, "upgrade_config.ps1 Updated",
         "Open config\\upgrade_config.ps1 and confirm all <-- UPDATE THIS fields are filled in correctly.",
         "Any machine", "Pending"),
    ]

    r = 4
    alt = False
    for num, check, how, machines, result in checks:
        bg = CLR["light_gray"] if alt else CLR["white"]
        wc(ws, r, 2, num, bold=True, bg=bg, h_align="center", border=True)
        wc(ws, r, 3, check, bold=True, bg=bg, border=True)
        wc(ws, r, 4, how, bg=bg, border=True)
        wc(ws, r, 5, machines, bg=bg, h_align="center", border=True)
        wc(ws, r, 6, result, bold=True, fg=CLR["amber"], bg=CLR["amber_bg"],
           h_align="center", border=True)
        ws.row_dimensions[r].height = 36
        r += 1
        alt = not alt

    # Footer note
    ws.row_dimensions[r].height = 28
    ws.merge_cells(f"B{r}:F{r}")
    wc(ws, r, 2,
       "✔  Proceed to LRE Server Upgrade sheet only when ALL checks are PASS on all machines.",
       bold=True, fg=CLR["green"], bg=CLR["green_bg"], sz=10, border=True)


# ==============================================================================
# SHEET 4 — LRE SERVER UPGRADE
# ==============================================================================
def build_server_upgrade(wb):
    ws = wb.create_sheet("LRE Server Upgrade")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 8), ("C", 28), ("D", 54),
                        ("E", 20), ("F", 18), ("G", 4)])

    ws.merge_cells("B1:F1")
    c = ws["B1"]
    c.value     = "LRE SERVER UPGRADE  |  Script: 01_Upgrade_LREServer.ps1"
    c.fill      = fill(CLR["navy"])
    c.font      = Font(bold=True, color=CLR["yellow_hdr"], size=14, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    col_headers(ws, 2, ["", "Step", "Action", "Detail / Command", "Note / Expected Result", "Status"],
                bg=CLR["navy"])
    ws.row_dimensions[2].height = 22
    freeze(ws, "B3")

    steps = [
        # Phase 1
        ("PHASE 1 — PRE-UPGRADE CHECKS  (Automated)", None, None, None),
        ("1.1", "Confirm installer share is accessible",
         "Script automatically checks that \\\\fileserver\\LRE_26.1 is reachable and setup_server.exe is found.",
         "Script exits with ERROR if share or EXE not found. Fix path in upgrade_config.ps1.", "Auto"),
        ("1.2", "Verify current version = 25.1",
         "Script reads Windows registry and confirms installed version matches 25.1.",
         "Script exits with ERROR if version doesn't match. Do not bypass.", "Auto"),
        ("1.3", "Check disk space ≥ 20 GB",
         "Script checks free space on C:\\. Fails if < 20 GB free.",
         "Free up disk space before re-running if this fails.", "Auto"),
        ("1.4", "Check no pending reboot",
         "Script checks pending reboot registry keys.",
         "Reboot the machine and re-run script if this fails.", "Auto"),

        # Phase 2
        ("PHASE 2 — GATHER CREDENTIALS  (Interactive prompt)", None, None, None),
        ("2.1", "Enter DB Admin password",
         "Script prompts:  Enter DB Admin password for 'sa'",
         "Type securely at the prompt. Leave blank in config to always prompt.", "Manual"),
        ("2.2", "Enter DB User password",
         "Script prompts:  Enter DB User password for 'lre_user'",
         "Type securely at the prompt.", "Manual"),
        ("2.3", "Enter System User password",
         "Script prompts:  Enter System User password for 'IUSR_METRO'",
         "Must match password for the Windows service account.", "Manual"),
        ("2.4", "Enter Secure Passphrase",
         "Script prompts:  Enter Secure Communication Passphrase (min 12 chars)",
         "Must be ≥ 12 alphanumeric characters. Note it down — must be the same on ALL nodes.", "Manual"),

        # Phase 3
        ("PHASE 3 — DATABASE BACKUP CONFIRMATION  (Manual confirm)", None, None, None),
        ("3.1", "Confirm database backups",
         "Script displays DB server/database names and asks:  'I confirm that database backups are complete. Continue?'",
         "Type  Y  to proceed. If DBA has not confirmed backups, type  N  and wait for confirmation.", "Manual"),

        # Phase 4
        ("PHASE 4 — STOP SERVICES  (Automated)", None, None, None),
        ("4.1", "Stop IIS",
         "Script runs:  iisreset /stop",
         "IIS (W3SVC) is stopped. LRE web UI becomes unavailable — expected.", "Auto"),
        ("4.2", "Stop LRE Backend Service",
         "Script stops LREBackend / OpenTextLREBackend service.",
         "Service names checked in order. Warning is logged if service not found (may already be stopped).", "Auto"),
        ("4.3", "Stop LRE Alerts Service",
         "Script stops LREAlerts / OpenTextLREAlerts service.",
         "Warning if not found — not critical.", "Auto"),
        ("4.4", "[CLUSTERED ONLY] Stop services on ALL other nodes",
         "If you have multiple LRE Servers (cluster): manually stop IIS + Backend + Alerts on EVERY other node BEFORE running the installer on any node.",
         "Failure to do this in a cluster will cause data corruption.", "Manual (if clustered)"),

        # Phase 5
        ("PHASE 5 — INSTALL PREREQUISITES  (Automated)", None, None, None),
        ("5.1", "Install .NET Framework 4.8",
         "Script runs ndp48-x86-x64-allos-enu.exe /q /norestart  (if not already installed).",
         "Exit codes 0, 3010, 1641 = success. Exit 3010 = reboot needed after install.", "Auto"),
        ("5.2", "Install .NET Core Hosting 8.x",
         "Script runs dotnet-hosting-8.0.17-win.exe /quiet",
         "Provides ASP.NET Core runtime required by LRE 26.1 backend.", "Auto"),
        ("5.3", "Install VC++ Redistributable x86 & x64",
         "Script runs vc_redist.x86.exe and vc_redist.x64.exe /quiet /norestart",
         "Required C++ runtime libraries for LRE components.", "Auto"),

        # Phase 6
        ("PHASE 6 — PREPARE UserInput.xml  (Automated)", None, None, None),
        ("6.1", "Generate UserInput.xml",
         "Script copies vendor's UserInput.xml from the share and merges your config values into it (DB server, credentials, passphrase, cert, paths).",
         "A timestamped XML is written to C:\\LRE_UpgradeLogs\\UserInput\\. Passwords are NOT logged.", "Auto"),

        # Phase 7
        ("PHASE 7 — RUN LRE SERVER INSTALLER  (Automated with confirm)", None, None, None),
        ("7.1", "Confirm installer launch",
         "Script asks:  'About to run LRE Server upgrade from setup_server.exe. Continue?'",
         "Type  Y  to launch. This starts the in-place upgrade — the existing 25.1 installation is upgraded to 26.1.", "Manual"),
        ("7.2", "Silent installer runs (up to 90 min)",
         "setup_server.exe /s USER_CONFIG_FILE_PATH=\"...xml\" INSTALLDIR=\"C:\\Program Files\\OpenText\\LRE\"",
         "DO NOT interrupt. A separate installer log is written to C:\\LRE_UpgradeLogs\\. Expected exit code: 0 (success) or 3010 (reboot needed).", "Auto"),
        ("7.3", "Check exit code",
         "0 = Success. 3010 = Reboot required. 1602 = Cancelled. 1603 = Fatal error (check installer log).",
         "If exit code is 1603 or unexpected: check C:\\LRE_UpgradeLogs\\LREServer_Installer_*.log for details before retrying.", "Auto"),

        # Phase 8
        ("PHASE 8 — POST-INSTALL VERIFICATION  (Automated)", None, None, None),
        ("8.1", "Read Configuration Wizard log",
         "Script reads last 30 lines of configurationWizardLog_pcs.txt from the install directory.",
         r"Logged to C:\LRE_UpgradeLogs\. Review if ERROR / FAIL / exception keywords are found.", "Auto"),

        # Phase 9
        ("PHASE 9 — START SERVICES  (Automated)", None, None, None),
        ("9.1", "Start IIS",
         "Script runs:  iisreset /start",
         "Waits 10 seconds for full IIS initialisation.", "Auto"),
        ("9.2", "Start LRE Backend Service",
         "Script starts LREBackend / OpenTextLREBackend.",
         "Service start confirmed in log.", "Auto"),
        ("9.3", "Start LRE Alerts Service",
         "Script starts LREAlerts / OpenTextLREAlerts.",
         "Service start confirmed in log.", "Auto"),

        # Phase 10
        ("PHASE 10 — CAPTURE PUBLIC KEY  (Automated)", None, None, None),
        ("10.1", "Wait 30 seconds for services to initialise",
         "Script pauses 30 s before querying the REST API.",
         "Allow time for LRE backend to fully start.", "Auto"),
        ("10.2", "Retrieve Public Key via REST API",
         "Script calls:  GET http://localhost/Admin/rest/v1/configuration/getPublicKey",
         "If successful, key is written to \\\\fileserver\\LRE_26.1\\PublicKey.txt and to C:\\LRE_UpgradeLogs\\PublicKey.txt.", "Auto"),
        ("10.3", "Fallback — read from pcs.config",
         r"If REST fails: script reads key from C:\Program Files\OpenText\LRE\dat\pcs.config",
         "If both methods fail: retrieve manually from the URL above or the Configuration Wizard 'Finish' page.", "Auto"),
        ("10.4", "[IF MANUAL] Provide Public Key to host scripts",
         "If key could not be auto-captured: set PublicKeyValue in upgrade_config.ps1 manually before running scripts 02–04.",
         "The Public Key is required by Controller, Data Processor, and Load Generator upgrade scripts.", "Manual (if needed)"),

        # Phase 11
        ("PHASE 11 — CONFIG CONSISTENCY & VERSION CHECK  (Automated)", None, None, None),
        ("11.1", "Verify pcs.config and appsettings.json",
         r"Script checks C:\Program Files\OpenText\LRE\dat\pcs.config and C:\LRE_Repository\system_config\appsettings.json for correct SystemUserName.",
         "WARN logged if user doesn't match. Review manually if WARNING.", "Auto"),
        ("11.2", "Confirm installed version = 26.1",
         "Script re-reads registry and confirms new version is 26.1.",
         "WARN if version doesn't match expected. Log the actual version.", "Auto"),
        ("11.3", "Clean up temp UserInput.xml",
         "Script deletes C:\\LRE_UpgradeLogs\\UserInput\\Server_UserInput_*.xml (contains passwords).",
         "Confirm file is deleted. If not deleted automatically, delete manually.", "Auto"),

        # Done
        ("COMPLETE — LRE Server Upgrade Finished", None, None, None),
        ("✔", "Verify LRE UI is accessible",
         "Open browser:  https://lre.yourdomain.com/LRE/",
         "LRE login page must load. If not: check IIS and backend service status.", "Manual"),
        ("✔", "Proceed to Host Upgrades sheet",
         "Run 02_Upgrade_Controller.ps1, then 03_Upgrade_DataProcessor.ps1, then 04_Upgrade_LoadGenerator.ps1 on respective machines.",
         "Do NOT skip scripts. Each host depends on the Public Key written by script 01.", "Next Step"),
    ]

    STATUS_COLORS = {
        "Auto":               (CLR["light_blue"], CLR["navy"]),
        "Manual":             (CLR["amber_bg"],   CLR["amber"]),
        "Manual (if needed)": (CLR["amber_bg"],   CLR["amber"]),
        "Manual (if clustered)": (CLR["red_bg"],  CLR["red"]),
        "Next Step":          (CLR["green_bg"],   CLR["green"]),
    }

    r = 3
    alt = False
    for item in steps:
        if item[1] is None:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            section_hdr(ws, r, 2, 6, f"  {item[0]}")
            ws.row_dimensions[r].height = 20
            r += 1
            alt = False
            continue

        step, action, detail, note, status = item
        bg = CLR["light_gray"] if alt else CLR["white"]
        s_bg, s_fg = STATUS_COLORS.get(status, (CLR["amber_bg"], CLR["amber"]))

        wc(ws, r, 2, step, bold=True, bg=bg, h_align="center", border=True, sz=9)
        wc(ws, r, 3, action, bold=True, bg=bg, border=True)
        wc(ws, r, 4, detail, bg=bg, border=True)
        wc(ws, r, 5, note, bg=bg, border=True, italic=True, fg=CLR["gray"], sz=9)
        wc(ws, r, 6, status, bold=True, fg=s_fg, bg=s_bg, h_align="center", border=True)

        ws.row_dimensions[r].height = 38
        r += 1
        alt = not alt


# ==============================================================================
# SHEET 5 — HOST UPGRADES
# ==============================================================================
def build_host_upgrades(wb):
    ws = wb.create_sheet("Host Upgrades")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 8), ("C", 22), ("D", 28),
                        ("E", 42), ("F", 22), ("G", 18), ("H", 4)])

    ws.merge_cells("B1:G1")
    c = ws["B1"]
    c.value     = "HOST UPGRADES  |  Scripts: 02 (Controller)  ·  03 (Data Processor)  ·  04 (Load Generator)"
    c.fill      = fill(CLR["navy"])
    c.font      = Font(bold=True, color=CLR["yellow_hdr"], size=14, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("B2:G2")
    wc(ws, 2, 2,
       "▶  Prerequisite: 01_Upgrade_LREServer.ps1 must have completed successfully and PublicKey.txt must be written to the installer share.",
       bold=True, fg=CLR["red"], bg=CLR["red_bg"], sz=10, border=True)
    ws.row_dimensions[2].height = 22

    col_headers(ws, 3,
                ["", "Step", "Host Type", "Action", "Detail / Command", "Machine(s)", "Status"],
                bg=CLR["navy"])
    ws.row_dimensions[3].height = 22
    freeze(ws, "B4")

    steps = [
        # Controller
        ("CONTROLLER  —  02_Upgrade_Controller.ps1", None, None, None, None),
        ("C-1", "Controller", "Pre-flight checks",
         "Script checks installer share, confirms current version = 25.1, disk space, pending reboot.",
         "Each Controller host", "Pending"),
        ("C-2", "Controller", "Enter credentials at prompt",
         "Prompts for: DB Admin password, DB User password, System User password, Secure Passphrase, Public Key (if not in config).",
         "Each Controller host", "Pending"),
        ("C-3", "Controller", "Stop services (auto)",
         "Script stops LoadRunner Agent Service and Remote Management Agent on this host.",
         "Each Controller host", "Pending"),
        ("C-4", "Controller", "Install prerequisites (auto)",
         "Installs .NET 4.8, .NET Core Hosting 8.x, VC++ Redist x86/x64 if not already present.",
         "Each Controller host", "Pending"),
        ("C-5", "Controller", "Prepare UserInput.xml (auto)",
         r"Merges config values into vendor's HostUserInput.xml. Written to C:\LRE_UpgradeLogs\UserInput\.",
         "Each Controller host", "Pending"),
        ("C-6", "Controller", "Run host installer (auto ~60 min)",
         "setup_host.exe /s USER_CONFIG_FILE_PATH=\"...xml\" INSTALLDIR=\"C:\\Program Files\\OpenText\\Performance Center Host\"",
         "Each Controller host", "Pending"),
        ("C-7", "Controller", "Start services (auto)",
         "Script restarts LoadRunner Agent Service and Remote Management Agent after successful install.",
         "Each Controller host", "Pending"),
        ("C-8", "Controller", "Verify version = 26.1",
         "Registry check confirms new version. Log is written to C:\\LRE_UpgradeLogs\\Controller_Upgrade_*.log.",
         "Each Controller host", "Pending"),

        # Data Processor
        ("DATA PROCESSOR  —  03_Upgrade_DataProcessor.ps1", None, None, None, None),
        ("D-1", "Data Processor", "Pre-flight checks",
         "Same as Controller: share access, version check (25.1), disk space, pending reboot.",
         "Each DP host", "Pending"),
        ("D-2", "Data Processor", "Enter credentials at prompt",
         "Same credential prompts as Controller.",
         "Each DP host", "Pending"),
        ("D-3", "Data Processor", "Stop services (auto)",
         "Stops Agent and Remote Mgmt Agent.",
         "Each DP host", "Pending"),
        ("D-4", "Data Processor", "Install prerequisites (auto)",
         "Same prerequisite set as Controller.",
         "Each DP host", "Pending"),
        ("D-5", "Data Processor", "Run host installer (auto ~60 min)",
         "setup_host.exe /s with DP-specific UserInput.xml values.",
         "Each DP host", "Pending"),
        ("D-6", "Data Processor", "Start services & verify (auto)",
         "Agent and Remote Mgmt started. Version check to 26.1. Log written.",
         "Each DP host", "Pending"),

        # Load Generator
        ("LOAD GENERATOR  —  04_Upgrade_LoadGenerator.ps1", None, None, None, None),
        ("L-1", "Load Generator", "Pre-flight checks",
         "Checks share, confirms OneLG version = 25.1, disk space, pending reboot.",
         "Each LG host", "Pending"),
        ("L-2", "Load Generator", "Determine installer method",
         "Script auto-detects: uses SetupOneLG.exe if found, falls back to OneLG_x64.msi.",
         "Each LG host", "Pending"),
        ("L-3", "Load Generator", "Enter credentials at prompt",
         "Prompts for System User password and Secure Passphrase.",
         "Each LG host", "Pending"),
        ("L-4", "Load Generator", "Stop LoadRunner Agent Service (auto)",
         "Stops magentservice / LoadRunnerAgentService before install.",
         "Each LG host", "Pending"),
        ("L-5", "Load Generator", "Install prerequisites (auto)",
         "Installs .NET 4.8, .NET Core Hosting, VC++ Redist.",
         "Each LG host", "Pending"),
        ("L-6", "Load Generator", "Run OneLG installer (auto ~60 min)",
         "SetupOneLG.exe /s with UserInput.xml  OR  msiexec /i OneLG_x64.msi /qn with properties.",
         "Each LG host", "Pending"),
        ("L-7", "Load Generator", "Start Agent & verify (auto)",
         "Restarts LoadRunner Agent Service. Verifies version = 26.1.",
         "Each LG host", "Pending"),

        # Final
        ("ALL HOSTS COMPLETE — Run Post-Upgrade Verification", None, None, None, None),
        ("✔", "All", "Proceed to Post-Upgrade Verify sheet",
         "Run 05_PostUpgradeVerify.ps1 on every machine to confirm full environment health.",
         "All machines", "Next Step"),
    ]

    STATUS_COLORS = {
        "Pending":   (CLR["amber_bg"], CLR["amber"]),
        "Next Step": (CLR["green_bg"], CLR["green"]),
    }

    r = 4
    alt = False
    for item in steps:
        if item[1] is None:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            section_hdr(ws, r, 2, 7, f"  {item[0]}")
            ws.row_dimensions[r].height = 20
            r += 1
            alt = False
            continue

        step, htype, action, detail, machines, status = item
        bg = CLR["light_gray"] if alt else CLR["white"]
        s_bg, s_fg = STATUS_COLORS.get(status, (CLR["amber_bg"], CLR["amber"]))

        wc(ws, r, 2, step, bold=True, bg=bg, h_align="center", border=True, sz=9)
        wc(ws, r, 3, htype, bold=True, bg=CLR["light_blue"],
           h_align="center", border=True, sz=9)
        wc(ws, r, 4, action, bold=True, bg=bg, border=True)
        wc(ws, r, 5, detail, bg=bg, border=True)
        wc(ws, r, 6, machines, bg=bg, h_align="center", border=True, italic=True,
           fg=CLR["gray"], sz=9)
        wc(ws, r, 7, status, bold=True, fg=s_fg, bg=s_bg, h_align="center", border=True)

        ws.row_dimensions[r].height = 36
        r += 1
        alt = not alt


# ==============================================================================
# SHEET 6 — POST-UPGRADE VERIFICATION
# ==============================================================================
def build_post_verify(wb):
    ws = wb.create_sheet("Post-Upgrade Verify")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 6), ("C", 32), ("D", 50),
                        ("E", 20), ("F", 18), ("G", 4)])

    ws.merge_cells("B1:F1")
    c = ws["B1"]
    c.value     = "POST-UPGRADE VERIFICATION  |  Script: 05_PostUpgradeVerify.ps1"
    c.fill      = fill(CLR["navy"])
    c.font      = Font(bold=True, color=CLR["yellow_hdr"], size=14, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("B2:F2")
    wc(ws, 2, 2,
       "▶  Run on EACH machine:  .\\05_PostUpgradeVerify.ps1   (elevated PowerShell — auto-detects Server / Controller / DP / LG)",
       bold=True, fg=CLR["teal"], bg=CLR["teal_light"], sz=10, border=True)
    ws.row_dimensions[2].height = 22

    col_headers(ws, 3, ["", "#", "Verification Check", "Expected Result / How To Verify",
                         "Machine(s)", "Result"], bg=CLR["navy"])
    ws.row_dimensions[3].height = 22
    freeze(ws, "B4")

    checks = [
        ("AUTOMATED — Script 05_PostUpgradeVerify.ps1  (runs all checks below)", None, None, None),
        (1,  "Version = 26.1",
         "Registry shows installed version 26.1. Script reports PASS.",
         "All machines", "Pending"),
        (2,  "IIS (W3SVC) Running",
         "Service status = Running. Script reports PASS.",
         "LRE Server", "Pending"),
        (3,  "LRE Backend Service Running",
         "LREBackend / OpenTextLREBackend = Running.",
         "LRE Server", "Pending"),
        (4,  "LRE Alerts Service Running",
         "LREAlerts / OpenTextLREAlerts = Running.",
         "LRE Server", "Pending"),
        (5,  "LoadRunner Agent Service Running",
         "magentservice / LoadRunnerAgentService = Running.",
         "Controller, DP, LG", "Pending"),
        (6,  "Remote Management Agent Running",
         "AlAgent / RemoteManagementAgent = Running.",
         "Controller, DP, LG", "Pending"),
        (7,  "LRE Web Application Accessible",
         "Script tests https://<FQDN>/LRE/ — expects HTTP 200 / 302.",
         "LRE Server", "Pending"),
        (8,  "pcs.config User Consistency",
         r"C:\Program Files\OpenText\LRE\dat\pcs.config contains SystemUserName (IUSR_METRO).",
         "LRE Server", "Pending"),
        (9,  "appsettings.json User Consistency",
         r"C:\LRE_Repository\system_config\appsettings.json contains SystemUserName.",
         "LRE Server", "Pending"),
        (10, "lts.config User Consistency",
         r"C:\Program Files\OpenText\Performance Center Host\dat\lts.config contains SystemUserName.",
         "Controller, DP", "Pending"),
        (11, "OneLG Config File Exists",
         r"C:\Program Files\OpenText\OneLG\dat\lts.config (or br_lnch_server.cfg) found.",
         "Load Generator", "Pending"),
        (12, "No Pending Reboot",
         "No pending reboot registry keys. Reboot and re-verify if FAIL.",
         "All machines", "Pending"),
        (13, "Temp UserInput.xml Files Cleaned Up",
         r"C:\LRE_UpgradeLogs\UserInput\ contains no *.xml files with credentials.",
         "All machines", "Pending"),
        (14, "No Recent LRE Event Log Errors",
         "No LRE/OpenText/LoadRunner Error events in Application log in last 2 hours.",
         "All machines", "Pending"),

        ("MANUAL CHECKS — Browser & Functional", None, None, None),
        (15, "Log in to LRE UI",
         "Open https://lre.yourdomain.com/LRE/ in browser. Log in as admin. Confirm landing page loads.",
         "LRE Server", "Pending"),
        (16, "Check Administration > About",
         "Verify version shown = 26.1 in the LRE Administration module.",
         "LRE Server", "Pending"),
        (17, "Verify all host machines appear in Administration",
         "In LRE Administration > Hosts, confirm Controllers, DPs, and LGs are listed and show 'Connected'.",
         "LRE Server (browser)", "Pending"),
        (18, "Run a smoke test scenario",
         "Run a short / known-good test scenario to confirm LRE can allocate hosts, run a load test, and process results.",
         "LRE Server (browser)", "Pending"),
        (19, "Confirm Public Key file on share",
         r"Check \\fileserver\LRE_26.1\PublicKey.txt exists and is non-empty.",
         "Any machine", "Pending"),
        (20, "Archive upgrade logs",
         r"Copy C:\LRE_UpgradeLogs\ from every machine to a central archive location.",
         "All machines", "Pending"),
        (21, "Close Change Request",
         "Update ITSM Change Request with completion status, evidence (log paths), and any deviations from this SOE.",
         "Change Manager", "Pending"),
    ]

    r = 4
    alt = False
    for item in checks:
        if item[1] is None:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            section_hdr(ws, r, 2, 6, f"  {item[0]}")
            ws.row_dimensions[r].height = 20
            r += 1
            alt = False
            continue

        num, check, expected, machines, result = item
        bg = CLR["light_gray"] if alt else CLR["white"]
        wc(ws, r, 2, num, bold=True, bg=bg, h_align="center", border=True)
        wc(ws, r, 3, check, bold=True, bg=bg, border=True)
        wc(ws, r, 4, expected, bg=bg, border=True)
        wc(ws, r, 5, machines, bg=bg, h_align="center", border=True,
           italic=True, fg=CLR["gray"], sz=9)
        wc(ws, r, 6, result, bold=True, fg=CLR["amber"], bg=CLR["amber_bg"],
           h_align="center", border=True)
        ws.row_dimensions[r].height = 36
        r += 1
        alt = not alt

    ws.row_dimensions[r].height = 28
    ws.merge_cells(f"B{r}:F{r}")
    wc(ws, r, 2,
       "✔  All checks PASS + smoke test complete = Upgrade Successful. Update the Change Request and close.",
       bold=True, fg=CLR["green"], bg=CLR["green_bg"], sz=10, border=True)


# ==============================================================================
# SHEET 7 — BACKOUT PLAN
# ==============================================================================
def build_backout(wb):
    ws = wb.create_sheet("Backout Plan")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [("A", 4), ("B", 8), ("C", 26), ("D", 56),
                        ("E", 22), ("F", 18), ("G", 4)])

    ws.merge_cells("B1:F1")
    c = ws["B1"]
    c.value     = "BACKOUT PLAN  —  LRE 25.1 → 26.1 Upgrade Rollback"
    c.fill      = fill(CLR["red"])
    c.font      = Font(bold=True, color=CLR["white"], size=14, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Key decision box
    ws.merge_cells("B2:F4")
    c = ws["B2"]
    c.value = (
        "WHEN TO INVOKE THIS PLAN:\n"
        "  • LRE Server installer exits with code 1603 (Fatal Error) AND cannot be resolved within the maintenance window.\n"
        "  • Post-install services fail to start after multiple attempts.\n"
        "  • LRE web UI is inaccessible after upgrade and the root cause cannot be identified quickly.\n"
        "  • Business impact is unacceptable and rolling forward is not feasible within the window.\n\n"
        "DECISION AUTHORITY:  The Change Manager (or designated technical lead) must formally declare 'Rollback' before executing these steps."
    )
    c.fill      = fill(CLR["red_bg"])
    c.font      = Font(bold=False, color=CLR["red"], size=10, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.border    = medium_border()
    ws.row_dimensions[2].height = 80

    col_headers(ws, 5, ["", "Step", "Action", "Detail / Command", "Machine(s)", "Owner"],
                bg=CLR["red"])
    ws.row_dimensions[5].height = 22
    freeze(ws, "B6")

    steps = [
        ("IMMEDIATE ACTIONS — Stop the upgrade", None, None, None),
        ("B-1", "Declare Rollback",
         "Change Manager verbally declares rollback. All team members stop any in-progress installer steps.",
         "All", "Change Manager"),
        ("B-2", "Do NOT reboot any machine",
         "Avoid reboots until the rollback plan is underway. Reboots during a partial install may leave machines unbootable.",
         "All", "All team"),
        ("B-3", "Document the failure point",
         "Note exactly which Phase/Step failed, the exit code, and the last lines of the installer log: C:\\LRE_UpgradeLogs\\LREServer_Installer_*.log",
         "LRE Server", "Tech Lead"),

        ("STOP ALL SERVICES ON ALL NODES", None, None, None),
        ("B-4", "Stop LRE Services (Server)",
         "Run:  iisreset /stop   then stop LREBackend and LREAlerts services.\n"
         "  net stop LREBackend\n  net stop LREAlerts",
         "LRE Server", "Tech Lead"),
        ("B-5", "Stop LRE Services (Hosts)",
         "On each Controller, DP, LG:\n"
         "  net stop magentservice\n  net stop AlAgent",
         "All hosts", "Tech Lead"),

        ("RESTORE DATABASE FROM BACKUP", None, None, None),
        ("B-6", "Contact DBA — restore databases",
         "DBA restores LRE_LAB, LRE_ADMIN, LRE_SITE from the pre-upgrade backup taken in Phase 3.",
         "SQL Server", "DBA"),
        ("B-7", "Confirm database restore complete",
         "DBA confirms all three databases are restored and accessible at the same connection strings (DbServerHost:DbServerPort).",
         "SQL Server", "DBA"),

        ("UNINSTALL 26.1 — RESTORE 25.1", None, None, None),
        ("B-8", "Uninstall LRE 26.1 from LRE Server",
         "Option A (preferred): Use Windows Add/Remove Programs to uninstall LRE 26.1.\n"
         "Option B: Run the 26.1 installer in repair/remove mode if available.\n"
         "Option C: msiexec /x {ProductCode} /qn  (find product code in registry: HKLM\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Uninstall)",
         "LRE Server", "Tech Lead"),
        ("B-9", "Reinstall LRE 25.1 from original media",
         "Run the 25.1 setup_server.exe with the original 25.1 UserInput.xml (retrieve from your pre-upgrade archive or backup).",
         "LRE Server", "Tech Lead"),
        ("B-10", "Uninstall 26.1 from Host machines (if upgraded)",
         "Repeat uninstall on any Controller/DP/LG that completed the 26.1 upgrade before rollback was declared.\n"
         "Use Add/Remove Programs or msiexec /x.",
         "Upgraded hosts", "Tech Lead"),
        ("B-11", "Reinstall LRE 25.1 on Host machines",
         "Run the original 25.1 setup_host.exe / SetupOneLG.exe on each host that was uninstalled in B-10.",
         "Upgraded hosts", "Tech Lead"),

        ("RESTORE CONFIGURATION FILES", None, None, None),
        ("B-12", "Restore pcs.config from backup",
         r"Copy backed-up pcs.config to C:\Program Files\OpenText\LRE\dat\pcs.config on the LRE Server.",
         "LRE Server", "Tech Lead"),
        ("B-13", "Restore appsettings.json from backup",
         r"Copy backed-up appsettings.json to C:\LRE_Repository\system_config\appsettings.json.",
         "LRE Server", "Tech Lead"),
        ("B-14", "Restore lts.config on hosts (if applicable)",
         r"Copy pre-upgrade lts.config to C:\Program Files\OpenText\Performance Center Host\dat\ on each host.",
         "Controller, DP", "Tech Lead"),

        ("RESTART SERVICES AND VALIDATE", None, None, None),
        ("B-15", "Start LRE Services (Server)",
         "Run:  iisreset /start\n  net start LREBackend\n  net start LREAlerts",
         "LRE Server", "Tech Lead"),
        ("B-16", "Start LRE Services (Hosts)",
         "On each host:\n  net start magentservice\n  net start AlAgent",
         "All hosts", "Tech Lead"),
        ("B-17", "Verify LRE UI is accessible on 25.1",
         "Open https://lre.yourdomain.com/LRE/ and confirm:\n  - Login page loads\n  - Administration > About shows version 25.1\n  - All hosts appear as Connected",
         "LRE Server", "Tech Lead"),
        ("B-18", "Run smoke test on 25.1",
         "Execute a known-good short test scenario to confirm LRE is fully functional at 25.1.",
         "LRE Server", "Test Lead"),

        ("COMMUNICATION & CLOSE-OUT", None, None, None),
        ("B-19", "Notify stakeholders",
         "Inform all stakeholders that the upgrade was rolled back and LRE is restored to 25.1.",
         "All", "Change Manager"),
        ("B-20", "Update Change Request",
         "Update ITSM Change Request with:\n  - Failure point and reason\n  - Rollback steps taken\n  - Current state (25.1 restored)\n  - Log file paths\n  - Next planned upgrade window",
         "All", "Change Manager"),
        ("B-21", "Conduct Post-Incident Review",
         "Schedule a PIR (Post-Incident Review) within 5 business days. Review failure root cause, adjust SOE if needed.",
         "All", "Change Manager"),
    ]

    r = 6
    alt = False
    for item in steps:
        if item[1] is None:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            bg = CLR["red"] if "IMMEDIATE" in item[0] else CLR["blue"]
            section_hdr(ws, r, 2, 6, f"  {item[0]}", bg=bg)
            ws.row_dimensions[r].height = 20
            r += 1
            alt = False
            continue

        step, action, detail, machines, owner = item
        bg = CLR["light_gray"] if alt else CLR["white"]

        wc(ws, r, 2, step, bold=True, bg=CLR["red_bg"], h_align="center", border=True,
           fg=CLR["red"], sz=9)
        wc(ws, r, 3, action, bold=True, bg=bg, border=True)
        wc(ws, r, 4, detail, bg=bg, border=True)
        wc(ws, r, 5, machines, bg=bg, h_align="center", border=True,
           italic=True, fg=CLR["gray"], sz=9)
        wc(ws, r, 6, owner, bold=True, bg=CLR["amber_bg"], fg=CLR["amber"],
           h_align="center", border=True, sz=9)

        ws.row_dimensions[r].height = 40
        r += 1
        alt = not alt

    ws.row_dimensions[r].height = 28
    ws.merge_cells(f"B{r}:F{r}")
    wc(ws, r, 2,
       "⚠  Rollback complete when: LRE 25.1 UI is accessible, smoke test passes, DBA confirms databases restored, Change Request updated.",
       bold=True, fg=CLR["amber"], bg=CLR["amber_bg"], sz=10, border=True)


# ==============================================================================
# MAIN
# ==============================================================================
def main():
    wb = openpyxl.Workbook()

    build_cover(wb)
    build_prerequisites(wb)
    build_pre_checks(wb)
    build_server_upgrade(wb)
    build_host_upgrades(wb)
    build_post_verify(wb)
    build_backout(wb)

    # Tab colours
    tab_colors = {
        "Cover":               "1F3864",
        "Prerequisites":       "2E75B6",
        "Pre-Upgrade Checks":  "375623",
        "LRE Server Upgrade":  "7030A0",
        "Host Upgrades":       "2E75B6",
        "Post-Upgrade Verify": "375623",
        "Backout Plan":        "9C0006",
    }
    for sheet_name, color in tab_colors.items():
        if sheet_name in wb.sheetnames:
            wb[sheet_name].sheet_properties.tabColor = color

    out_path = r"c:\Workspace\LRE-Automates-Scripts\LRE_Upgrade_SOE.xlsx"
    wb.save(out_path)
    print(f"SOE workbook saved: {out_path}")


if __name__ == "__main__":
    main()
